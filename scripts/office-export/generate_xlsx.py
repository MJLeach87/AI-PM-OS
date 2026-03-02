"""
XLSX generator for PM OS Office Export.
Converts a Markdown metrics/data-analysis file to a branded Excel workbook
using openpyxl.

Template-aware: maps PM OS data-analysis section headers to named sheets.
Generic fallback: one sheet per H2 when no template match is detected.

Branding spec: pm-os-reference/identity/BRANDING.md
Constants:      scripts/office-export/branding.py
"""

import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import branding as B


# ---------------------------------------------------------------------------
# Known PM OS data-analysis section → sheet name mapping
# ---------------------------------------------------------------------------
SECTION_SHEET_MAP = {
    'executive summary':    'Executive Summary',
    'metric review':        'Metric Review',
    'instrumentation gaps': 'Instrumentation Gaps',
    'data quality':         'Data Quality',
    'sql library':          'SQL Library',
    'a/b test setup':       'AB Test Setup',
    'ab test setup':        'AB Test Setup',
    'recommendations':      'Recommendations',
}


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _header_fill() -> PatternFill:
    return PatternFill('solid', fgColor=B.hex_to_openpyxl_color(B.PRIMARY))


def _alt_fill() -> PatternFill:
    return PatternFill('solid', fgColor=B.hex_to_openpyxl_color(B.SURFACE_ALT))


def _surface_fill() -> PatternFill:
    return PatternFill('solid', fgColor=B.hex_to_openpyxl_color(B.SURFACE))


def _accent_fills(accent: B.AccentPair):
    return (
        PatternFill('solid', fgColor=B.hex_to_openpyxl_color(accent.bg)),
        Font(name=B.FONT_PRIMARY, size=B.FONT_SIZE_BODY,
             color=B.hex_to_openpyxl_color(accent.text)),
    )


def _thin_border() -> Border:
    side = Side(style='thin', color=B.BORDER.lstrip('#'))
    return Border(left=side, right=side, top=side, bottom=side)


def _style_header_cell(cell, text: str):
    cell.value = text
    cell.font = Font(
        name=B.FONT_PRIMARY, size=B.FONT_SIZE_BODY,
        bold=True, color='FFFFFFFF'
    )
    cell.fill = _header_fill()
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    cell.border = _thin_border()


def _style_data_cell(cell, text: str, row_idx: int):
    cell.value = text
    bg_fill = _surface_fill() if row_idx % 2 == 0 else _alt_fill()
    cell.fill = bg_fill
    cell.font = Font(
        name=B.FONT_PRIMARY, size=B.FONT_SIZE_BODY,
        color=B.hex_to_openpyxl_color(B.TEXT_SECONDARY)
    )
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = _thin_border()


def _status_style(cell, text: str, row_idx: int):
    """Apply semantic accent colour to a status cell if text matches a known token."""
    lower = text.strip().lower()
    accent = None
    if lower in ('active', 'complete', 'done', 'shipped'):
        accent = B.ACCENTS.prd
    elif lower in ('draft', 'in progress', 'wip'):
        accent = B.ACCENTS.draft
    elif lower in ('superseded', 'archived', 'removed'):
        accent = B.ACCENTS.superseded
    elif lower in ('ost',):
        accent = B.ACCENTS.ost
    elif lower in ('tech spec', 'technical spec'):
        accent = B.ACCENTS.tech_spec
    elif lower in ('gtm',):
        accent = B.ACCENTS.gtm
    elif lower in ('prototype',):
        accent = B.ACCENTS.prototype

    if accent:
        fill, font = _accent_fills(accent)
        cell.fill = fill
        cell.font = font
    else:
        _style_data_cell(cell, text, row_idx)

    cell.value = text
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = _thin_border()


def _auto_col_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or '')
                # Multi-line: use longest line
                longest = max((len(line) for line in val.split('\n')), default=0)
                max_len = max(max_len, longest)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 60)


# ---------------------------------------------------------------------------
# Markdown section parser
# ---------------------------------------------------------------------------

def _split_into_sections(md_text: str) -> list[dict]:
    """
    Split Markdown into sections keyed by H2 heading.
    Returns list of {'heading': str, 'lines': [str]}.
    """
    sections = []
    current = {'heading': 'Overview', 'lines': []}

    for line in md_text.splitlines():
        m = re.match(r'^##\s+(.*)', line)
        if m:
            if current['lines'] or current['heading'] != 'Overview':
                sections.append(current)
            current = {'heading': m.group(1).strip(), 'lines': []}
        else:
            current['lines'].append(line)

    if current['lines'] or current['heading'] != 'Overview':
        sections.append(current)

    return sections


def _extract_table(lines: list[str]) -> tuple[list[str], list[list[str]]]:
    """Extract the first Markdown table from a list of lines."""
    table_lines = [l for l in lines if l.strip().startswith('|')]
    if len(table_lines) < 2:
        return [], []
    header = [c.strip() for c in table_lines[0].split('|') if c.strip()]
    rows = []
    for tl in table_lines[2:]:
        row = [c.strip() for c in tl.split('|') if c.strip()]
        if row:
            rows.append(row)
    return header, rows


def _write_section_to_sheet(ws, section: dict):
    """
    Write a section's content to a worksheet.
    If the section contains a Markdown table, render it as a styled table.
    Otherwise write the raw text lines.
    """
    header, rows = _extract_table(section['lines'])

    if header:
        # Write table header
        for col_idx, col_name in enumerate(header, start=1):
            _style_header_cell(ws.cell(row=1, column=col_idx), col_name)

        # Write data rows
        status_cols = {i for i, h in enumerate(header, start=1)
                       if 'status' in h.lower() or 'type' in h.lower()}
        for row_num, row_data in enumerate(rows, start=2):
            for col_idx in range(1, len(header) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                val = row_data[col_idx - 1] if col_idx <= len(row_data) else ''
                if col_idx in status_cols:
                    _status_style(cell, val, row_num - 2)
                else:
                    _style_data_cell(cell, val, row_num - 2)

        # Freeze top row
        ws.freeze_panes = 'A2'

    else:
        # Plain text fallback — write non-empty lines as single-cell rows
        row_num = 1
        for line in section['lines']:
            stripped = line.strip()
            if not stripped:
                continue
            # Strip leading Markdown markers
            stripped = re.sub(r'^#+\s+', '', stripped)
            stripped = re.sub(r'^[\-\*\+]\s+', '', stripped)
            stripped = re.sub(r'\*\*(.+?)\*\*', r'\1', stripped)
            stripped = re.sub(r'`(.+?)`', r'\1', stripped)
            cell = ws.cell(row=row_num, column=1, value=stripped)
            cell.font = Font(name=B.FONT_PRIMARY, size=B.FONT_SIZE_BODY,
                             color=B.hex_to_openpyxl_color(B.TEXT_SECONDARY))
            cell.alignment = Alignment(wrap_text=True)
            row_num += 1

    _auto_col_width(ws)


# ---------------------------------------------------------------------------
# Workbook builder
# ---------------------------------------------------------------------------

def generate(source_path: Path, out_dir: Path) -> Path:
    """Convert a Markdown file to a branded XLSX workbook. Returns output path."""
    md_text = source_path.read_text(encoding='utf-8')
    sections = _split_into_sections(md_text)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default empty sheet

    for section in sections:
        heading_lower = section['heading'].lower()
        sheet_name = SECTION_SHEET_MAP.get(heading_lower, section['heading'][:31])
        # Excel sheet names: max 31 chars, no special chars
        sheet_name = re.sub(r'[\\/:*?\[\]]', '', sheet_name)[:31]

        ws = wb.create_sheet(title=sheet_name)
        _write_section_to_sheet(ws, section)

    stem = source_path.stem
    out_path = out_dir / f'{stem}.xlsx'
    wb.save(str(out_path))
    return out_path
